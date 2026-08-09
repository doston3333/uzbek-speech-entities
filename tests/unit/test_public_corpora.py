"""Offline unit coverage for public-corpus import and speech-year adapters."""

from __future__ import annotations

import hashlib
import json
from pathlib import Path

import openpyxl
import pytest

from uzbek_speech_entities.ner.labels import validate_bio_record
from uzbek_speech_entities.ner.public_corpora import (
    COMMON_VOICE_UZ,
    NEWS_YOUTUBE_UZ,
    UZBEKVOICE_UZ,
    SpeechCorpus,
    calendar_year_to_spoken_words,
    convert_bioes_to_bio,
    extract_numeric_calendar_years,
    extract_spoken_calendar_years,
    import_uzner_workbook,
    iter_parquet_row_group_text,
    merge_spoken_year_record_batches,
    mine_spoken_year_records,
    normalize_uzbek_token,
    normalized_phrase,
    parse_spoken_year,
    verify_hf_revision,
)


def _workbook(path: Path) -> str:
    workbook = openpyxl.Workbook()
    experts = workbook.active
    experts.title = "Experts"
    experts.append(("Sentence", "Token", "GOLD-TAG"))
    experts.append((1, "Ali", "B-PER"))
    experts.append((1, "keldi", "O"))
    dataset_one = workbook.create_sheet("Dataset_1")
    dataset_one.append(("Sentence", "Token", "BIOES-Tag"))
    dataset_one.append((2, "ikki", "B-DATE"))
    dataset_one.append((2, "ming", "I-DATE"))
    dataset_one.append((2, "yilda", "E-DATE"))
    dataset_one.append((3, "Toshkent", "S-GPE"))
    dataset_two = workbook.create_sheet("Dataset_2")
    dataset_two.append(("Sentence", "Token", "BIOES-Tag"))
    dataset_two.append((4, "besh", "S-TIME"))
    workbook.save(path)
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _read_jsonl(path: Path) -> list[dict[str, object]]:
    return [json.loads(line) for line in path.read_text(encoding="utf-8").splitlines()]


def test_bioes_conversion_maps_to_strict_canonical_bio() -> None:
    labels = convert_bioes_to_bio(
        ["Janubiy", "Toshkent", "va", "kitob"],
        ["B-GPE", "E-GPE", "O", "S-WORK_OF_ART"],
    )
    assert labels == ["B-LOC", "I-LOC", "O", "B-WORK"]
    validate_bio_record({"id": "strict", "tokens": ["a", "b", "c", "d"], "ner_tags": labels})


def test_unsupported_source_label_rejects_whole_sentence() -> None:
    with pytest.raises(ValueError, match="unsupported entity"):
        convert_bioes_to_bio(["x", "y"], ["S-ALIEN", "O"])


def test_noisy_uzner_types_strip_to_o_while_keeping_core_entities() -> None:
    labels = convert_bioes_to_bio(
        ["Ali", "va", "Mahsulot", "hodisa"],
        ["S-PER", "O", "S-PRODUCT", "S-EVENT"],
    )
    assert labels == ["B-PER", "O", "O", "O"]
    with pytest.raises(ValueError, match="unsupported entity"):
        convert_bioes_to_bio(["x"], ["S-ALIEN"])


def test_expert_and_temporal_workbook_modes_are_separate(tmp_path: Path) -> None:
    workbook = tmp_path / "uzner.xlsx"
    expected_hash = _workbook(workbook)
    expert_output, expert_stats = tmp_path / "expert.jsonl", tmp_path / "expert-stats.json"
    import_uzner_workbook(
        workbook, expert_output, expert_stats, mode="expert", expected_sha256=expected_hash
    )
    assert _read_jsonl(expert_output)[0]["ner_tags"] == ["B-PER", "O"]

    excluded = tmp_path / "exclude.jsonl"
    excluded.write_text('{"tokens":["besh"]}\n', encoding="utf-8")
    temporal_output, temporal_stats = tmp_path / "temporal.jsonl", tmp_path / "temporal-stats.json"
    stats = import_uzner_workbook(
        workbook,
        temporal_output,
        temporal_stats,
        mode="temporal",
        exclude_jsonl_paths=[excluded],
        expected_sha256=expected_hash,
    )
    temporal_records = _read_jsonl(temporal_output)
    assert [record["tokens"] for record in temporal_records] == [["ikki", "ming", "yilda"]]
    assert temporal_records[0]["ner_tags"] == ["B-TEMPORAL", "I-TEMPORAL", "I-TEMPORAL"]
    assert stats["source"] == {
        "doi": "10.17632/48923w3gyr.1",
        "license": "CC BY 4.0",
        "sha256": expected_hash,
        "url": "https://data.mendeley.com/public-files/datasets/48923w3gyr/files/"
        "bf0fb867-8b70-40ff-893e-64ab5ab78cfd/file_downloaded",
    }


def test_apostrophe_normalization_and_spoken_year_parser() -> None:
    assert normalize_uzbek_token("TOʻQQIZ") == "to'qqiz"
    assert normalized_phrase(["toʻqqiz"]) == normalized_phrase(["to'qqiz"])
    assert parse_spoken_year(["ikki", "ming", "va", "besh"]) == 2005
    candidates = extract_spoken_calendar_years("ikki ming va besh yilni reja qildik")
    assert candidates == [(["ikki", "ming", "va", "besh", "yilni"], 2005, "yilni")]


def test_strict_digit_year_normalization_emits_spoken_words_only() -> None:
    assert calendar_year_to_spoken_words(1998) == (
        "bir",
        "ming",
        "toʻqqiz",
        "yuz",
        "toʻqson",
        "sakkiz",
    )
    assert extract_numeric_calendar_years("2026-yilda uchrashamiz") == [
        (["ikki", "ming", "yigirma", "olti", "yilda"], 2026, "yilda")
    ]
    assert extract_numeric_calendar_years("2026 yil davomida ishladik") == []
    assert extract_numeric_calendar_years("1799 yil va 2036 yil") == []
    assert all(
        parse_spoken_year(calendar_year_to_spoken_words(year)) == year for year in range(1800, 2036)
    )


@pytest.mark.parametrize(
    "text",
    ("ikki yuz yil", "ikki ming yil davomida", "ikki ming so'm berdim"),
)
def test_year_miner_rejects_duration_and_non_year_cardinals(text: str) -> None:
    assert extract_spoken_calendar_years(text) == []


def test_mining_deduplicates_deterministically_and_emits_phrase_only_bio() -> None:
    corpus = SpeechCorpus("example/corpus", "0" * 40, "train", "text", "CC0")
    records, stats = mine_spoken_year_records(
        corpus,
        [
            "ikki ming va besh yilni",
            "bir ming toʻqqiz yuz toʻqson sakkiz yili",
            "ikki ming va besh yilni",
        ],
    )
    assert [record["tokens"] for record in records] == [
        ["bir", "ming", "toʻqqiz", "yuz", "toʻqson", "sakkiz", "yili"],
        ["ikki", "ming", "va", "besh", "yilni"],
    ]
    assert stats["connector_counts"] == {"va": 1}
    assert stats["suffix_counts"] == {"yili": 1, "yilni": 1}
    assert stats["transformation_counts"] == {"observed_spoken": 2}
    for record in records:
        validate_bio_record(record)
        assert all(tag.endswith("TEMPORAL") for tag in record["ner_tags"])


def test_digit_year_records_are_normalized_with_auditable_provenance() -> None:
    corpus = SpeechCorpus("example/corpus", "0" * 40, "train", "text", "CC0")
    records, stats = mine_spoken_year_records(corpus, ["2026-yil boshlandi"])
    assert records == [
        {
            "id": records[0]["id"],
            "tokens": ["ikki", "ming", "yigirma", "olti", "yil"],
            "ner_tags": [
                "B-TEMPORAL",
                "I-TEMPORAL",
                "I-TEMPORAL",
                "I-TEMPORAL",
                "I-TEMPORAL",
            ],
            "source": f"hf:example/corpus@{'0' * 40}",
            "augmentation": {
                "source_text_form": "2026",
                "transformation": "digit_calendar_year_to_spoken",
            },
        }
    ]
    assert stats["transformation_counts"] == {"digit_calendar_year_to_spoken": 1}


def test_shard_merge_preserves_transformations_and_prefers_observed_speech() -> None:
    corpus = SpeechCorpus("example/corpus", "0" * 40, "train", "text", "CC0")
    digit_records, _ = mine_spoken_year_records(
        corpus, ["2025-yil boshlandi", "2026-yil boshlandi"]
    )
    observed_records, _ = mine_spoken_year_records(corpus, ["ikki ming yigirma olti yil boshlandi"])
    records, stats = merge_spoken_year_record_batches(corpus, [digit_records, observed_records])
    by_value = {parse_spoken_year(record["tokens"][:-1]): record for record in records}
    assert by_value[2025]["augmentation"] == {
        "source_text_form": "2025",
        "transformation": "digit_calendar_year_to_spoken",
    }
    assert by_value[2026]["augmentation"] is None
    assert stats["transformation_counts"] == {
        "digit_calendar_year_to_spoken": 1,
        "observed_spoken": 1,
    }


def test_allowlist_revision_is_enforced_without_network(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(
        "uzbek_speech_entities.ner.public_corpora._response_json",
        lambda _url: {"sha": "not-the-pinned-revision"},
    )
    with pytest.raises(ValueError, match="revision mismatch"):
        verify_hf_revision(COMMON_VOICE_UZ)


def test_news_youtube_source_is_immutably_pinned_and_licensed() -> None:
    assert NEWS_YOUTUBE_UZ.dataset_id == "islomov/news_youtube_uzbek_speech_dataset"
    assert NEWS_YOUTUBE_UZ.revision == "bbff3fb27cbf461260f2b5f93e5f56d0c4008a6c"
    assert NEWS_YOUTUBE_UZ.text_column == "text"
    assert NEWS_YOUTUBE_UZ.license == "Apache-2.0"


def test_uzbekvoice_source_is_immutably_pinned_and_licensed() -> None:
    assert UZBEKVOICE_UZ.dataset_id == "ai4uz/uzbekvoice-filtered"
    assert UZBEKVOICE_UZ.revision == "b392eae07f28911b1538215c130bf056f7b2f7fa"
    assert UZBEKVOICE_UZ.text_column == "sentence"
    assert UZBEKVOICE_UZ.license == "Apache-2.0"


def test_parquet_row_group_iterator_requests_only_text_column() -> None:
    calls: list[tuple[int, list[str], bool]] = []

    class Column:
        def to_pylist(self) -> list[object]:
            return ["ikki ming besh yil", None]

    class Table:
        def column(self, name: str) -> Column:
            assert name == "text"
            return Column()

    class FakeParquet:
        class metadata:  # noqa: N801 - matches PyArrow's external attribute.
            num_row_groups = 2

        def read_row_group(self, index: int, *, columns: list[str], use_threads: bool) -> Table:
            calls.append((index, columns, use_threads))
            return Table()

    assert list(iter_parquet_row_group_text(FakeParquet(), "text")) == [
        "ikki ming besh yil",
        "ikki ming besh yil",
    ]
    assert calls == [(0, ["text"], False), (1, ["text"], False)]
