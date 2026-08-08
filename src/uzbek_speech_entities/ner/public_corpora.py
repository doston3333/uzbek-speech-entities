"""Deterministic adapters for the pinned public Uzbek NER and speech corpora.

The functions in this module deliberately keep downloads and dataset traversal out
of training code.  They emit only the project's canonical prepared-record shape.
"""

from __future__ import annotations

import hashlib
import json
import re
import tempfile
import unicodedata
from collections import Counter
from collections.abc import Iterable, Iterator, Mapping, Sequence
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from urllib.parse import quote
from urllib.request import urlopen

from .labels import BIOValidationError, validate_bio_record

UZNER_DOI = "10.17632/48923w3gyr.1"
UZNER_URL = (
    "https://data.mendeley.com/public-files/datasets/48923w3gyr/files/"
    "bf0fb867-8b70-40ff-893e-64ab5ab78cfd/file_downloaded"
)
UZNER_SHA256 = "d0d50fa1dfb83cd66abf39076207968681e67ee80814301fa3248f256ff171d0"
UZNER_LICENSE = "CC BY 4.0"

ENTITY_MAP: dict[str, str] = {
    "PER": "PER",
    "PERSON": "PER",
    "ORG": "ORG",
    "GPE": "LOC",
    "LOC": "LOC",
    "FAC": "LOC",
    "DATE": "TEMPORAL",
    "TIME": "TEMPORAL",
    "DURATION": "TEMPORAL",
    "MONEY": "MONEY",
    "CARDINAL": "NUMERIC",
    "ORDINAL": "NUMERIC",
    "QUANTITY": "NUMERIC",
    "PERCENT": "NUMERIC",
    "AGE": "NUMERIC",
    "WORK_OF_ART": "WORK",
    "DOCUMENT": "WORK",
    "LAW": "WORK",
    "MISC": "MISC",
}

# Public UzNER types that would otherwise reject a whole sentence. Strip their
# spans to O so PER/ORG/LOC (and other mapped) labels in the same sentence are kept
# without injecting noisy MISC/ORG aliases that hurt general NER.
STRIP_TO_O_ENTITY_TYPES: frozenset[str] = frozenset(
    {
        "EVENT",
        "NORP",
        "POSITION",
        "PRODUCT",
        "LANGUAGE",
        "AWARD",
        "MEDIA",
        "PROGRAM",
        "VEHICLE",
        "COURSE",
        "PHONE",
        "URL",
    }
)

_BIOES_PREFIXES = frozenset(("B", "I", "O", "E", "S"))
_APOSTROPHES = "'`ʻʼʹʾˈ"
_APOSTROPHE_TRANSLATION: dict[str, str | int | None] = {
    character: "'" for character in _APOSTROPHES
}
_TOKEN_RE = re.compile(r"[^\W_]+(?:['`ʻʼʹʾˈ][^\W_]+)*", flags=re.UNICODE)
_YEAR_SUFFIXES = frozenset(("yil", "yili", "yilda", "yilni", "yilga", "yildan"))
_YEAR_WORDS = {
    "bir": 1,
    "ikki": 2,
    "uch": 3,
    "tort": 4,
    "besh": 5,
    "olti": 6,
    "yetti": 7,
    "sakkiz": 8,
    "toqqiz": 9,
    "on": 10,
    "yigirma": 20,
    "ottiz": 30,
    "qirq": 40,
    "ellik": 50,
    "oltmish": 60,
    "yetmish": 70,
    "sakson": 80,
    "toqson": 90,
}
_YEAR_WORD_SURFACES = {
    1: "bir",
    2: "ikki",
    3: "uch",
    4: "toʻrt",
    5: "besh",
    6: "olti",
    7: "yetti",
    8: "sakkiz",
    9: "toʻqqiz",
    10: "oʻn",
    20: "yigirma",
    30: "oʻttiz",
    40: "qirq",
    50: "ellik",
    60: "oltmish",
    70: "yetmish",
    80: "sakson",
    90: "toʻqson",
}
_DURATION_FOLLOWERS = frozenset(("davomida", "mobaynida", "ichida", "oldin", "keyin"))


@dataclass(frozen=True)
class SpeechCorpus:
    """An explicitly allowlisted Hugging Face speech source."""

    dataset_id: str
    revision: str
    split: str
    text_column: str
    license: str
    hf_config: str | None = None

    @property
    def dataset_url(self) -> str:
        """Stable human-readable URL for the pinned public dataset revision."""
        return f"https://huggingface.co/datasets/{self.dataset_id}/tree/{self.revision}"


COMMON_VOICE_UZ = SpeechCorpus(
    dataset_id="yakhyo/mozilla-common-voice-uzbek",
    revision="09f89fbf98a7d73a394ae80921950966a5569c1c",
    split="train",
    text_column="text",
    license="CC0 (claimed inherited)",
)
USC_UZ = SpeechCorpus(
    dataset_id="murodbek/uzbek-speech-corpus",
    revision="257f7e46f0a92d81ba00f22659ec93213a3b5f7e",
    split="train",
    text_column="sentence",
    license="CC BY 4.0",
)
NEWS_YOUTUBE_UZ = SpeechCorpus(
    dataset_id="islomov/news_youtube_uzbek_speech_dataset",
    revision="bbff3fb27cbf461260f2b5f93e5f56d0c4008a6c",
    split="train",
    text_column="text",
    license="Apache-2.0",
)
IT_YOUTUBE_UZ = SpeechCorpus(
    dataset_id="islomov/it_youtube_uzbek_speech_dataset",
    revision="1d4b0e37b489a66e59ee363a44f5a4ac2900458b",
    split="train",
    text_column="text",
    license="Apache-2.0",
)
PODCASTS_TASHKENT_UZ = SpeechCorpus(
    dataset_id="islomov/podcasts_tashkent_dialect_youtube_uzbek_speech_dataset",
    revision="a397215c80771174cdc63ef83dce79bf8d6c06fd",
    split="train",
    text_column="text",
    license="Apache-2.0",
)
UZBEKVOICE_UZ = SpeechCorpus(
    dataset_id="ai4uz/uzbekvoice-filtered",
    revision="b392eae07f28911b1538215c130bf056f7b2f7fa",
    split="train",
    text_column="sentence",
    license="Apache-2.0",
)
UZBEKVOICE2_UZ = SpeechCorpus(
    dataset_id="Jurabek/uzbekvoice-filtered2",
    revision="ab2b460ac96c01659a4c816664fd9bd860946957",
    split="train",
    text_column="sentence",
    license="Apache-2.0",
)
UZBEK_NEWS_TEXT = SpeechCorpus(
    dataset_id="MLDataScientist/Uzbek_news_dataset",
    revision="cf0008591608b3ff8cd2df7eb6e1ae70d4e932c4",
    split="train",
    text_column="text",
    license="CC BY 4.0",
)
WIKIPEDIA_UZ = SpeechCorpus(
    dataset_id="MLDataScientist/Wikipedia-uzbek-2024-05-01",
    revision="c4884cd832a9f8e8fe3c56543d12614fb6d3e491",
    split="train",
    text_column="text",
    license="Apache-2.0",
)
FLEURS_UZ = SpeechCorpus(
    dataset_id="google/fleurs",
    revision="70bb2e84b976b7e960aa89f1c648e09c59f894dd",
    split="train",
    text_column="transcription",
    license="CC BY 4.0",
    hf_config="uz_uz",
)
GLOTCC_UZN_LATN = SpeechCorpus(
    dataset_id="cis-lmu/GlotCC-V1",
    revision="9ad140b6be3ac7b539606a2b4809b49d122823de",
    split="train",
    text_column="content",
    license="CC0-1.0",
    hf_config="uzn-Latn",
)
LEGAL_UZ = SpeechCorpus(
    dataset_id="sukhrobnurali/uzbek-legal-corpus-v1",
    revision="834614ce82bc700820e002d3a896070a4d2b59de",
    split="train",
    text_column="article_text",
    license="Apache-2.0",
    hf_config="articles",
)
ZERO_SHOT_NEWS_UZ = SpeechCorpus(
    dataset_id="risqaliyevds/uzbek-zero-shot-classification",
    revision="a4ad4607a7c1ffe492fe1420a05fb6c7b6165383",
    split="train",
    text_column="text",
    license="MIT",
)
OASST2_UZ = SpeechCorpus(
    dataset_id="MLDataScientist/oasst2_uzbek",
    revision="67e8d30915d943735bdee6b44018c4f891ff794e",
    split="train",
    text_column="text",
    license="Apache-2.0",
)
SPEECH_CORPORA: tuple[SpeechCorpus, ...] = (
    COMMON_VOICE_UZ,
    USC_UZ,
    NEWS_YOUTUBE_UZ,
    IT_YOUTUBE_UZ,
    PODCASTS_TASHKENT_UZ,
    UZBEKVOICE_UZ,
    UZBEKVOICE2_UZ,
    UZBEK_NEWS_TEXT,
    WIKIPEDIA_UZ,
    FLEURS_UZ,
    GLOTCC_UZN_LATN,
    ZERO_SHOT_NEWS_UZ,
    OASST2_UZ,
)


def sha256_file(path: Path) -> str:
    """Return a streaming SHA-256 digest for ``path``."""
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def normalize_uzbek_token(token: str) -> str:
    """Normalize case and apostrophe variants without transliterating Uzbek text."""
    normalized = unicodedata.normalize("NFKC", token).lower()
    return normalized.translate(str.maketrans(_APOSTROPHE_TRANSLATION))


def _word_key(token: str) -> str:
    return normalize_uzbek_token(token).replace("'", "")


def tokenize_uzbek(text: str) -> list[str]:
    """Split speech text into word tokens while retaining apostrophe-bearing words."""
    return _TOKEN_RE.findall(unicodedata.normalize("NFKC", text))


def normalized_phrase(tokens: Sequence[str]) -> str:
    """Stable normalization used only for duplicate detection."""
    return " ".join(_word_key(token) for token in tokens)


def _split_source_tag(tag: object) -> tuple[str, str | None]:
    if not isinstance(tag, str):
        raise ValueError(f"source tag must be a string, got {tag!r}")
    clean = tag.strip().upper()
    if clean == "O":
        return ("O", None)
    if "-" not in clean:
        raise ValueError(f"malformed source tag: {tag!r}")
    prefix, entity_type = clean.split("-", maxsplit=1)
    if prefix not in _BIOES_PREFIXES or prefix == "O" or not entity_type:
        raise ValueError(f"malformed source tag: {tag!r}")
    return (prefix, entity_type)


def convert_bioes_to_bio(tokens: Sequence[str], source_tags: Sequence[object]) -> list[str]:
    """Map supported BIOES tags to strict canonical BIO, rejecting whole bad sentences.

    Types in ``STRIP_TO_O_ENTITY_TYPES`` become ``O`` so sentences that also carry
    PER/ORG/LOC (etc.) remain usable without aliasing those spans into MISC/ORG.
    """
    if len(tokens) != len(source_tags) or not tokens:
        raise ValueError("tokens and source tags must be equally sized nonempty sequences")
    parsed = [_split_source_tag(tag) for tag in source_tags]
    unsupported = sorted(
        {
            entity
            for _, entity in parsed
            if entity
            and entity not in ENTITY_MAP
            and entity not in STRIP_TO_O_ENTITY_TYPES
        }
    )
    if unsupported:
        raise ValueError(f"unsupported entity type(s): {unsupported!r}")

    output: list[str] = []
    active: str | None = None
    for index, (prefix, source_type) in enumerate(parsed):
        if prefix == "O":
            output.append("O")
            active = None
            continue
        assert source_type is not None
        if source_type in STRIP_TO_O_ENTITY_TYPES:
            output.append("O")
            active = None
            continue
        entity_type = ENTITY_MAP[source_type]
        if prefix == "S":
            output.append(f"B-{entity_type}")
            active = None
        elif prefix == "B":
            output.append(f"B-{entity_type}")
            active = entity_type
        elif prefix == "I":
            if active != entity_type:
                raise ValueError(f"orphan I-{source_type} at token {index}")
            output.append(f"I-{entity_type}")
        elif prefix == "E":
            if active != entity_type:
                raise ValueError(f"orphan E-{source_type} at token {index}")
            output.append(f"I-{entity_type}")
            active = None
        else:  # pragma: no cover - _split_source_tag validates prefixes.
            raise ValueError(f"unsupported source prefix: {prefix!r}")
    if active is not None and any(prefix in {"E", "S"} for prefix, _ in parsed):
        raise ValueError(f"unterminated BIOES entity: {active}")
    try:
        validate_bio_record({"id": "validation", "tokens": list(tokens), "ner_tags": output})
    except BIOValidationError as error:
        raise ValueError(f"BIO conversion failed validation: {error}") from error
    return output


def _column_index(headers: Sequence[object], candidates: Sequence[str]) -> int:
    normalized = {str(value).strip().casefold(): index for index, value in enumerate(headers)}
    for candidate in candidates:
        index = normalized.get(candidate.casefold())
        if index is not None:
            return index
    raise ValueError(
        f"worksheet missing one of columns {list(candidates)!r}; got {list(headers)!r}"
    )


def _worksheet_sentences(
    sheet: Any, tag_column: str
) -> Iterator[tuple[str, list[str], list[object]]]:
    rows = sheet.iter_rows(values_only=True)
    try:
        headers = next(rows)
    except StopIteration:
        return
    sentence_index = _column_index(
        headers, ("Sentence", "SentenceID", "Sentence ID", "sentence_id")
    )
    token_index = _column_index(headers, ("Token", "Word", "Words", "token"))
    tag_index = _column_index(headers, (tag_column,))
    grouped: dict[str, tuple[list[str], list[object]]] = {}
    for row_number, row in enumerate(rows, start=2):
        if len(row) <= max(sentence_index, token_index, tag_index):
            raise ValueError(f"short row in {sheet.title!r} at row {row_number}")
        sentence = row[sentence_index]
        token = row[token_index]
        tag = row[tag_index]
        if sentence is None and token is None and tag is None:
            continue
        if sentence is None or token is None or tag is None:
            raise ValueError(f"incomplete row in {sheet.title!r} at row {row_number}")
        sentence_id = str(sentence).strip()
        token_text = str(token).strip()
        if not sentence_id or not token_text:
            raise ValueError(f"blank sentence or token in {sheet.title!r} at row {row_number}")
        bucket = grouped.setdefault(sentence_id, ([], []))
        bucket[0].append(token_text)
        bucket[1].append(tag)
    for sentence_id in sorted(grouped, key=lambda value: (len(value), value)):
        tokens, tags = grouped[sentence_id]
        yield sentence_id, tokens, tags


def _load_workbook(path: Path) -> Any:
    try:
        import openpyxl  # type: ignore[import-untyped]
    except ImportError as error:  # pragma: no cover - dependency error is environment-specific.
        raise RuntimeError("openpyxl is required for UzNER workbook import") from error
    return openpyxl.load_workbook(path, read_only=True, data_only=True)


def _read_excluded_phrases(paths: Sequence[Path]) -> set[str]:
    excluded: set[str] = set()
    for path in paths:
        with path.open(encoding="utf-8") as source:
            for line_number, line in enumerate(source, start=1):
                if not line.strip():
                    raise ValueError(f"blank JSONL line at {path}:{line_number}")
                value = json.loads(line)
                if not isinstance(value, Mapping):
                    raise ValueError(f"non-object JSONL value at {path}:{line_number}")
                tokens = value.get("tokens")
                if not isinstance(tokens, list) or not all(
                    isinstance(item, str) for item in tokens
                ):
                    raise ValueError(f"record lacks string tokens at {path}:{line_number}")
                excluded.add(normalized_phrase(tokens))
    return excluded


def _atomic_jsonl(path: Path, records: Iterable[Mapping[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary_name: str | None = None
    try:
        with tempfile.NamedTemporaryFile(
            mode="w", encoding="utf-8", dir=path.parent, prefix=f".{path.name}.", delete=False
        ) as temporary:
            temporary_name = temporary.name
            for record in records:
                temporary.write(
                    json.dumps(record, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
                )
                temporary.write("\n")
        Path(temporary_name).replace(path)
    except Exception:
        if temporary_name is not None:
            Path(temporary_name).unlink(missing_ok=True)
        raise


def _atomic_json(path: Path, value: Mapping[str, object]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary_name: str | None = None
    try:
        with tempfile.NamedTemporaryFile(
            mode="w", encoding="utf-8", dir=path.parent, prefix=f".{path.name}.", delete=False
        ) as temporary:
            temporary_name = temporary.name
            json.dump(value, temporary, ensure_ascii=False, sort_keys=True, indent=2)
            temporary.write("\n")
        Path(temporary_name).replace(path)
    except Exception:
        if temporary_name is not None:
            Path(temporary_name).unlink(missing_ok=True)
        raise


def import_uzner_workbook(
    workbook_path: Path,
    output_path: Path,
    statistics_path: Path,
    *,
    mode: str,
    exclude_jsonl_paths: Sequence[Path] = (),
    expected_sha256: str = UZNER_SHA256,
) -> dict[str, object]:
    """Convert a hash-pinned UzNER workbook into deterministic canonical JSONL."""
    if mode not in {"expert", "temporal"}:
        raise ValueError("mode must be 'expert' or 'temporal'")
    actual_hash = sha256_file(workbook_path)
    if actual_hash != expected_sha256:
        raise ValueError(
            f"workbook checksum mismatch: expected {expected_sha256}, got {actual_hash}"
        )
    sheet_names = (
        (("Experts", "GOLD-TAG"),)
        if mode == "expert"
        else (
            ("Dataset_1", "BIOES-Tag"),
            ("Dataset_2", "BIOES-Tag"),
        )
    )
    workbook = _load_workbook(workbook_path)
    excluded = _read_excluded_phrases(exclude_jsonl_paths)
    records: list[dict[str, object]] = []
    rejected: Counter[str] = Counter()
    label_counts: Counter[str] = Counter()
    try:
        for sheet_name, tag_column in sheet_names:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"workbook missing required sheet {sheet_name!r}")
            for sentence_id, tokens, tags in _worksheet_sentences(workbook[sheet_name], tag_column):
                try:
                    labels = convert_bioes_to_bio(tokens, tags)
                except ValueError as error:
                    rejected[str(error)] += 1
                    continue
                if mode == "temporal" and "B-TEMPORAL" not in labels:
                    rejected["no TEMPORAL entity"] += 1
                    continue
                if mode == "temporal" and normalized_phrase(tokens) in excluded:
                    rejected["excluded normalized duplicate"] += 1
                    continue
                record: dict[str, object] = {
                    "id": f"uzner-{mode}:{sheet_name}:{sentence_id}",
                    "tokens": tokens,
                    "ner_tags": labels,
                    "source": f"uzner:{UZNER_DOI}:{sheet_name}",
                    "augmentation": None,
                }
                validate_bio_record(record)
                records.append(record)
                label_counts.update(labels)
    finally:
        workbook.close()
    records.sort(key=lambda record: str(record["id"]))
    stats: dict[str, object] = {
        "accepted_records": len(records),
        "exclude_jsonl_paths": [str(path) for path in exclude_jsonl_paths],
        "label_counts": dict(sorted(label_counts.items())),
        "mode": mode,
        "rejected": dict(sorted(rejected.items())),
        "source": {
            "doi": UZNER_DOI,
            "license": UZNER_LICENSE,
            "sha256": actual_hash,
            "url": UZNER_URL,
        },
        "sheets": [name for name, _ in sheet_names],
    }
    _atomic_jsonl(output_path, records)
    _atomic_json(statistics_path, stats)
    return stats


def parse_spoken_year(words: Sequence[str]) -> int | None:
    """Parse a cardinal Uzbek year phrase, allowing ``va`` only between number words."""
    if not words:
        return None
    current = 0
    total = 0
    saw_number = False
    saw_thousand = False
    previous_was_connector = False
    for index, token in enumerate(words):
        key = _word_key(token)
        if key == "va":
            if index == 0 or index == len(words) - 1 or previous_was_connector:
                return None
            previous_was_connector = True
            continue
        previous_was_connector = False
        if key in _YEAR_WORDS:
            current += _YEAR_WORDS[key]
            saw_number = True
        elif key == "yuz":
            if saw_thousand and current == 0:
                return None
            current = (current or 1) * 100
            saw_number = True
        elif key == "ming":
            if saw_thousand:
                return None
            total += (current or 1) * 1000
            current = 0
            saw_thousand = True
            saw_number = True
        else:
            return None
    value = total + current
    return value if saw_number and 1800 <= value <= 2035 else None


def calendar_year_to_spoken_words(value: int) -> tuple[str, ...]:
    """Render one allowlisted calendar year as canonical Uzbek cardinal words."""
    if not 1800 <= value <= 2035:
        return ()
    words: list[str] = []
    thousands, remainder = divmod(value, 1000)
    words.extend((_YEAR_WORD_SURFACES[thousands], "ming"))
    hundreds, remainder = divmod(remainder, 100)
    if hundreds:
        words.extend((_YEAR_WORD_SURFACES[hundreds], "yuz"))
    tens, units = divmod(remainder, 10)
    if tens:
        words.append(_YEAR_WORD_SURFACES[tens * 10])
    if units:
        words.append(_YEAR_WORD_SURFACES[units])
    return tuple(words)


def extract_numeric_calendar_years(text: str) -> list[tuple[list[str], int, str]]:
    """Normalize strict digit-year + suffix spans from real speech transcripts."""
    tokens = tokenize_uzbek(text)
    candidates: list[tuple[list[str], int, str]] = []
    for suffix_index, suffix_token in enumerate(tokens):
        suffix = _word_key(suffix_token)
        if suffix not in _YEAR_SUFFIXES or suffix_index == 0:
            continue
        if (
            suffix_index + 1 < len(tokens)
            and _word_key(tokens[suffix_index + 1]) in _DURATION_FOLLOWERS
        ):
            continue
        raw_year = _word_key(tokens[suffix_index - 1])
        if not raw_year.isascii() or not raw_year.isdigit():
            continue
        words = calendar_year_to_spoken_words(int(raw_year))
        if words:
            candidates.append(([*words, suffix_token], int(raw_year), suffix))
    return candidates


def extract_spoken_calendar_years(text: str) -> list[tuple[list[str], int, str]]:
    """Extract high-precision number-word + calendar-year-suffix spans from speech text."""
    tokens = tokenize_uzbek(text)
    candidates: list[tuple[list[str], int, str]] = []
    for suffix_index, suffix_token in enumerate(tokens):
        suffix = _word_key(suffix_token)
        if suffix not in _YEAR_SUFFIXES:
            continue
        if (
            suffix_index + 1 < len(tokens)
            and _word_key(tokens[suffix_index + 1]) in _DURATION_FOLLOWERS
        ):
            continue
        start_limit = max(0, suffix_index - 7)
        valid: list[tuple[int, int]] = []
        for start in range(start_limit, suffix_index):
            span = tokens[start:suffix_index]
            value = parse_spoken_year(span)
            if value is not None:
                valid.append((start, value))
        if not valid:
            continue
        start, value = min(valid, key=lambda candidate: candidate[0])
        if start > 0 and _word_key(tokens[start - 1]) in set(_YEAR_WORDS) | {"yuz", "ming", "va"}:
            continue
        phrase = tokens[start : suffix_index + 1]
        candidates.append((phrase, value, suffix))
    return candidates


def mine_spoken_year_records(
    corpus: SpeechCorpus,
    texts: Iterable[str],
) -> tuple[list[dict[str, object]], dict[str, object]]:
    """Mine one source's transcript strings, emitting phrase-only TEMPORAL records."""
    records_by_phrase: dict[str, tuple[dict[str, object], int, str, str]] = {}
    for text in texts:
        if not isinstance(text, str):
            continue
        candidates = [
            *(phrase + ("observed_spoken",) for phrase in extract_spoken_calendar_years(text)),
            *(
                phrase + ("digit_calendar_year_to_spoken",)
                for phrase in extract_numeric_calendar_years(text)
            ),
        ]
        for phrase, value, suffix, transformation in candidates:
            key = normalized_phrase(phrase)
            existing = records_by_phrase.get(key)
            if existing is not None and (
                existing[3] == "observed_spoken" or transformation != "observed_spoken"
            ):
                continue
            digest = hashlib.sha256(key.encode("utf-8")).hexdigest()[:16]
            record: dict[str, object] = {
                "id": f"speech-year:{corpus.dataset_id}:{digest}",
                "tokens": phrase,
                "ner_tags": ["B-TEMPORAL", *("I-TEMPORAL" for _ in phrase[1:])],
                "source": f"hf:{corpus.dataset_id}@{corpus.revision}",
                "augmentation": (
                    None
                    if transformation == "observed_spoken"
                    else {
                        "source_text_form": str(value),
                        "transformation": transformation,
                    }
                ),
            }
            validate_bio_record(record)
            records_by_phrase[key] = (record, value, suffix, transformation)
    ordered = [records_by_phrase[key] for key in sorted(records_by_phrase)]
    records = [record for record, _, _, _ in ordered]
    return records, _spoken_year_statistics(corpus, ordered)


def _spoken_year_statistics(
    corpus: SpeechCorpus,
    entries: Sequence[tuple[dict[str, object], int, str, str]],
) -> dict[str, object]:
    return {
        "accepted_records": len(entries),
        "connector_counts": dict(
            sorted(
                Counter(
                    "va"
                    for record, _, _, _ in entries
                    if "va" in normalized_phrase(_record_tokens(record)).split()
                ).items()
            )
        ),
        "source": {
            "dataset_id": corpus.dataset_id,
            "dataset_url": corpus.dataset_url,
            "hf_config": corpus.hf_config,
            "license": corpus.license,
            "revision": corpus.revision,
            "split": corpus.split,
            "text_column": corpus.text_column,
        },
        "suffix_counts": dict(
            sorted(Counter(suffix for _, _, suffix, _ in entries).items())
        ),
        "transformation_counts": dict(
            sorted(Counter(kind for _, _, _, kind in entries).items())
        ),
        "value_counts": dict(
            sorted(Counter(str(value) for _, value, _, _ in entries).items())
        ),
    }


def _record_tokens(record: Mapping[str, object]) -> list[str]:
    """Return checked canonical tokens from an internally constructed record."""
    tokens = record["tokens"]
    if isinstance(tokens, list) and all(isinstance(token, str) for token in tokens):
        return tokens
    raise ValueError("canonical record has invalid tokens")


def _record_transformation(record: Mapping[str, object]) -> str:
    augmentation = record.get("augmentation")
    if augmentation is None:
        return "observed_spoken"
    if isinstance(augmentation, Mapping):
        transformation = augmentation.get("transformation")
        if transformation == "digit_calendar_year_to_spoken":
            return "digit_calendar_year_to_spoken"
    raise ValueError("speech-year record has invalid transformation provenance")


def merge_spoken_year_record_batches(
    corpus: SpeechCorpus,
    batches: Iterable[Sequence[Mapping[str, object]]],
) -> tuple[list[dict[str, object]], dict[str, object]]:
    """Deduplicate shard results while preserving their exact provenance metadata."""
    expected_source = f"hf:{corpus.dataset_id}@{corpus.revision}"
    records_by_phrase: dict[str, dict[str, object]] = {}
    for batch in batches:
        for value in batch:
            record = dict(value)
            validate_bio_record(record)
            if record.get("source") != expected_source:
                raise ValueError("speech-year shard record has mismatched source provenance")
            tokens = _record_tokens(record)
            if len(tokens) < 2 or _word_key(tokens[-1]) not in _YEAR_SUFFIXES:
                raise ValueError("speech-year shard record has an invalid year suffix")
            if parse_spoken_year(tokens[:-1]) is None:
                raise ValueError("speech-year shard record no longer parses as a calendar year")
            key = normalized_phrase(tokens)
            existing = records_by_phrase.get(key)
            if existing is None or (
                _record_transformation(existing) != "observed_spoken"
                and _record_transformation(record) == "observed_spoken"
            ):
                records_by_phrase[key] = record
    records = [records_by_phrase[key] for key in sorted(records_by_phrase)]
    entries = [
        (
            record,
            value,
            _word_key(tokens[-1]),
            _record_transformation(record),
        )
        for record in records
        for tokens in (_record_tokens(record),)
        for value in (parse_spoken_year(tokens[:-1]),)
        if value is not None
    ]
    return records, _spoken_year_statistics(corpus, entries)


def _response_json(url: str) -> object:
    with urlopen(url, timeout=30) as response:  # noqa: S310 - pinned public API URL.
        return json.loads(response.read().decode("utf-8"))


def verify_hf_revision(corpus: SpeechCorpus) -> None:
    """Require Hugging Face to resolve exactly the pinned commit before mining."""
    encoded_id = quote(corpus.dataset_id, safe="/")
    url = f"https://huggingface.co/api/datasets/{encoded_id}/revision/{corpus.revision}"
    payload = _response_json(url)
    if not isinstance(payload, Mapping) or payload.get("sha") != corpus.revision:
        actual = payload.get("sha") if isinstance(payload, Mapping) else None
        raise ValueError(
            f"Hugging Face revision mismatch for {corpus.dataset_id}: "
            f"expected {corpus.revision}, got {actual!r}"
        )


def discover_parquet_urls(corpus: SpeechCorpus) -> list[str]:
    """Discover only the pinned train Parquet shard URLs through datasets-server."""
    query = f"dataset={quote(corpus.dataset_id, safe='')}&revision={corpus.revision}"
    payload = _response_json(f"https://datasets-server.huggingface.co/parquet?{query}")
    if not isinstance(payload, Mapping) or not isinstance(payload.get("parquet_files"), list):
        raise ValueError("unexpected datasets-server parquet response")
    urls: list[str] = []
    for item in payload["parquet_files"]:
        if not isinstance(item, Mapping):
            continue
        if item.get("split") != corpus.split or not isinstance(item.get("url"), str):
            continue
        if corpus.hf_config is not None and item.get("config") != corpus.hf_config:
            continue
        urls.append(item["url"])
    if not urls:
        config_note = f" config={corpus.hf_config!r}" if corpus.hf_config else ""
        raise ValueError(
            f"no {corpus.split!r} parquet files found for {corpus.dataset_id}{config_note}"
        )
    return sorted(urls)


def iter_parquet_row_group_text(parquet_file: Any, text_column: str) -> Iterator[str]:
    """Yield one transcript column by row group, never asking Arrow for audio columns."""
    for row_group in range(parquet_file.metadata.num_row_groups):
        table = parquet_file.read_row_group(
            row_group,
            columns=[text_column],
            use_threads=False,
        )
        for value in table.column(text_column).to_pylist():
            if isinstance(value, str):
                yield value


def iter_parquet_text_column(urls: Sequence[str], text_column: str) -> Iterator[str]:
    """Range-read one transcript column with bounded retries; never materialize audio."""
    import fsspec  # type: ignore[import-untyped]
    import pyarrow.parquet as parquet  # type: ignore[import-untyped]

    for url in urls:
        last_error: OSError | None = None
        for _attempt in range(3):
            try:
                # HTTPFileSystem performs range requests.  The block cap prevents
                # prefetching a whole audio-backed shard while Arrow seeks footer
                # and the selected text column chunks.
                with fsspec.open(
                    url, "rb", block_size=8 * 1024 * 1024, timeout=30
                ).open() as source:
                    parquet_file = parquet.ParquetFile(source, pre_buffer=False)
                    yield from iter_parquet_row_group_text(parquet_file, text_column)
                last_error = None
                break
            except OSError as error:
                last_error = error
        if last_error is not None:
            raise RuntimeError(f"failed to range-read transcript parquet {url!r}") from last_error


def mine_allowlisted_speech(output_path: Path, statistics_path: Path) -> dict[str, object]:
    """Verify, discover, and mine every pinned allowlisted speech source deterministically."""
    all_records: dict[str, dict[str, object]] = {}
    source_stats: list[dict[str, object]] = []
    for corpus in SPEECH_CORPORA:
        verify_hf_revision(corpus)
        urls = discover_parquet_urls(corpus)
        records, stats = mine_spoken_year_records(
            corpus, iter_parquet_text_column(urls, corpus.text_column)
        )
        source_stats.append({**stats, "parquet_urls": urls})
        for record in records:
            key = normalized_phrase(_record_tokens(record))
            all_records.setdefault(key, record)
    records = [all_records[key] for key in sorted(all_records)]
    connector_counts: Counter[str] = Counter()
    suffix_counts: Counter[str] = Counter()
    value_counts: Counter[str] = Counter()
    source_counts: Counter[str] = Counter()
    for record in records:
        tokens = _record_tokens(record)
        if "va" in normalized_phrase(tokens).split():
            connector_counts["va"] += 1
        suffix_counts[_word_key(tokens[-1])] += 1
        value = parse_spoken_year(tokens[:-1])
        if value is None:  # pragma: no cover - records are created only by the parser above.
            raise ValueError("mined record no longer parses as a calendar year")
        value_counts[str(value)] += 1
        source = record["source"]
        if not isinstance(source, str):  # pragma: no cover - records are internally constructed.
            raise ValueError("canonical record has invalid source")
        source_counts[source] += 1
    statistics: dict[str, object] = {
        "accepted_records": len(records),
        "connector_counts": dict(sorted(connector_counts.items())),
        "deduplicated_by": "normalized_phrase",
        "source_counts": dict(sorted(source_counts.items())),
        "sources": source_stats,
        "suffix_counts": dict(sorted(suffix_counts.items())),
        "value_counts": dict(sorted(value_counts.items())),
    }
    _atomic_jsonl(output_path, records)
    _atomic_json(statistics_path, statistics)
    return statistics
